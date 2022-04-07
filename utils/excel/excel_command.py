import logging
# import os

# import django
import openpyxl


# def setup_django():
#     os.environ.setdefault(
#         "DJANGO_SETTINGS_MODULE",
#         "django_project.telegrambot.telegrambot.settings"
#     )
#     os.environ.update({'DJANGO_ALLOW_ASYNC_UNSAFE': "true"})
#     django.setup()
#
#
# setup_django()


from django_project.telegrambot.usersmanage.models import Item, Category


def delete_all_items():
    items = Item.objects.all()
    for item in items:
        item.delete()


def load_to_excel():
    wb = openpyxl.load_workbook('СписокТоваров.xlsx', read_only=True)
    ws = wb.active
    row = 3
    columns = ['B', 'C', 'D', 'E']
    for r in range(row, ws.max_row + 1):
        res_column = []
        category_name = None
        for c in columns:
            res_name = ws[f"{c}{r}"].value
            if res_name:
                res_column.append(str(res_name))
                if c == columns[2]:
                    category_name = str(res_name)
        employment = ws[f"B{r}"].value
        size = ws[f"C{r}"].value
        model = ws[f"D{r}"].value
        factory = ws[f"E{r}"].value

        category = None
        if category_name:
            category, create = Category.objects.get_or_create(name=category_name,
                                                              name_uz=category_name,
                                                              name_en=category_name)
        try:
            name = ' / '.join(res_column)
        except Exception as err:
            logging.info(err)
            break
        if name and category:
            try:
                Item(name=name, name_en=name, name_uz=name, category=category,
                     employment=employment, size=size, model=model, factory=factory).save()
            except Exception as err:
                print(err)


# if __name__ == '__main__':
#     load_to_excel()
#     # delete_all_items()
