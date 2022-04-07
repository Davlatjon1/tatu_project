import asyncio
import logging
from typing import Union
import datetime as dt

from django_project.telegrambot.constants.models import Constants
from handlers import state_data, botControlOrder
from loader import dp, bot, bot_tel
from aiogram import types
from aiogram.dispatcher import FSMContext
from aiogram.utils import markdown as md
from data import config
from django_project.telegrambot.usersmanage import models
from django_project.telegrambot.usersmanage.models import User
from keyboards.default import keyboards as kb
from keyboards.inline import inlines
from states.default import StateDefault
from utils.db_api import db_commands
from utils.misc import rate_limit
from utils.values import others, menu_values, values_title


async def main_menu(language: str, message: types.Message):
    await welcome_message(message=message, language=language)
    markup = await kb.main_keyboards(language)
    await choose_below(message=message, language=language, markup=markup)
    await search(language=language, message=message)
    await reset_data_user()
    await StateDefault.main.set()


async def home_state(language: str, message: types.Message, reset: bool = True):
    markup = await kb.main_keyboards(language=language)
    await choose_below(message=message, language=language, markup=markup)
    await StateDefault.main.set()
    if reset:
        await reset_data_user()


async def after_product(language: str, message: types.Message, user: User = None, all_product=False,
                        text: str = None, photo_or_video: bool = True):
    # if config.SEND_FILE_ITEMS:
    #     await send_file_items(message=message, language=language)
    #     return
    if user and all_product:
        await send_all_product(user=user, message=message, photo_or_video=photo_or_video)
        return

    if not text:
        text = values_title.CHOOSE_CATEGORY[language]
    markup = await kb.category_keyboards(language=language)
    await message.answer(text=text,
                         parse_mode=types.ParseMode.HTML,
                         reply_markup=markup)
    await StateDefault.category.set()


async def send_no_baskets(message: types.Message, language: str):
    await message.answer(values_title.EMPTY_BASKET[language])


async def send_all_product(user: User, message: types.Message, photo_or_video=True):
    items = await db_commands.get_item_filter(show_bot=True)
    i = 0
    for item in items:
        i = i + 1
        await send_item(item=item, user=user, message=message, photo_or_video=photo_or_video, save_data=False)
        if i % 29 == 0:
            # await asyncio.sleep(1)
            pass


async def send_all_product_text(message: types.Message, language: str):
    items = await db_commands.get_item_filter(show_bot=True, running=True)
    current_date = dt.datetime.now()
    text = f"<b>Дефицитный лист от {current_date.strftime('%d.%m.%Y')} на время {int(current_date.time().hour)}:{current_date.strftime('%M:%S')}</b>\n\n"
    items_text = others.items_text(items, language)
    text += items_text
    await botControlOrder.split_message(message=message, text=text, markup=None, give_markup=False)


async def after_basket(message: types.Message, language: str, user: User):
    basket_items = await db_commands.get_basket_item(user=user)
    if len(basket_items) == 0:
        await send_no_baskets(message=message, language=language)
        return
    basket_text = values_title.get_title_order(language=language, basket_items=basket_items)
    markup = await kb.basket_keyboards(language)

    basket_message, results = await botControlOrder.split_message(message=message, text=basket_text, markup=markup)
    await StateDefault.basket.set()

    async with dp.current_state().proxy() as data:
        basket_messages = data.get("basket_message")
        if type(basket_messages) == list:
            for basket_message in basket_messages:
                await basket_message.delete()

    await dp.current_state().update_data(basket_message=results)


async def show_orders(message: types.Message, user: User):
    basket_items = await db_commands.get_basket_item(user=user)
    language = user.language
    if len(basket_items) == 0:
        await send_no_baskets(message=message, language=language)
        return

    additional_order = await state_data.get_structure_from_more_order()
    basket_text = values_title.get_title_order(language=language, basket_items=basket_items,
                                               additional_order=additional_order)
    markup = await kb.basket_keyboards(language)
    basket_message, results = await botControlOrder.split_message(message=message, text=basket_text, markup=markup,
                                                                  give_markup=not additional_order['access'])
    async with dp.current_state().proxy() as data:
        basket_messages = data.get("basket_message")
        if type(basket_messages) == list:
            for basket_message in basket_messages:
                try:
                    await basket_message.delete()
                except:
                    pass
    message_search = await search(language=language, message=message)
    results.append(message_search)

    await dp.current_state().update_data(basket_message=results)
    if not additional_order['access']:
        await StateDefault.basket.set()


async def is_empty_basket(user: User, message: types.Message):
    count_basket = await db_commands.get_basket_item(user=user, count=True)
    is_empty = False
    if count_basket == 0:
        await send_no_baskets(message=message, language=user.language)
        await home_state(language=user.language, message=message)
        is_empty = True
    return is_empty


async def after_category(category: str, message: types.Message, language: str):
    sub_or_items = await menu_values.get_subcategories_or_items_by_category(language=language,
                                                                            category=category)

    markup = await kb.sub_or_item_keyboards_by_category(language=language, category=category, which='',
                                                        sub_or_items=sub_or_items.get(others.SUBCATEGORIES_OR_ITEMS))
    caption = values_title.get_caption_cat_or_product(language=language,
                                                      items=sub_or_items.get(others.ITEMS_CATEGORY),
                                                      subcategory=sub_or_items.get(others.SUBCATEGORIES))
    await message.answer(text=caption, reply_markup=markup)
    await StateDefault.subcategory.set()


async def after_subcategory(message: types.Message, language: str, items: Union[set, list] = None,
                            category: str = None, subcategory: str = None):
    markup = await kb.sub_or_item_keyboards_by_category(language=language, category=category,
                                                        subcategory=subcategory, sub_or_items=items,
                                                        which=others.ITEMS_FULL)
    await message.answer(text=values_title.CHOOSE_PRODUCT[language],
                         reply_markup=markup)
    await StateDefault.item.set()


async def after_item(item_name: str, message: types.Message, user: User):
    language = user.language
    attr = others.get_attribute_by_language(language=language)
    item = await db_commands.get_item_filter(first=True, **{f"{attr}": item_name})
    if item:
        await send_item(item=item, user=user, message=message)
    else:
        await message.answer(values_title.NOT_FOUND_ITEM[language])


async def choose_category(message: types.Message, language: str,
                          markup: Union[types.InlineKeyboardMarkup, types.ReplyKeyboardMarkup]):
    if markup:
        await message.answer(text=values_title.CHOOSE_CATEGORY[language], reply_markup=markup)
    else:
        await message.answer(text=values_title.CHOOSE_CATEGORY[language])


async def choose_below(language: str, message: types.Message, markup: Union[types.ReplyKeyboardMarkup,
                                                                            types.InlineKeyboardMarkup] = None):
    if markup:
        await message.answer(text=values_title.CHOOSE_BELOW[language], reply_markup=markup)
    else:
        await message.answer(text=values_title.CHOOSE_BELOW[language])


async def choose_product(language: str, message: types.Message, markup: Union[types.ReplyKeyboardMarkup,
                                                                              types.InlineKeyboardMarkup] = None):
    if markup:
        await message.answer(text=values_title.CHOOSE_PRODUCT[language], reply_markup=markup)
    else:
        await message.answer(text=values_title.CHOOSE_PRODUCT[language])


async def change_language(user_id: Union[str, int], user_lang, user: User = None):
    list_of_lang = list(menu_values.CHANGE_LANGUAGE.keys())
    index_find_lan = list_of_lang.index(user_lang)
    has_changed_lan = list_of_lang[(index_find_lan + 1) % len(list_of_lang)]
    await db_commands.change_language(user_id, has_changed_lan, user)
    return has_changed_lan


async def about_us(message: types.Message, language: str):
    abouts_us = await db_commands.get_abouts_us(language=language)
    if len(abouts_us) == 0:
        await message.answer(values_title.EMPTY[language])

    for au in abouts_us:
        text = ''
        if au.title:
            text += md.hbold(au.title)
        if au.description:
            text += '\n' + au.description
        text = await replace_text_special(message=message, text=text)
        if au.video_url_file_id:
            try:
                video_file_id = au.video_url_file_id
                await message.answer_video(video=video_file_id, caption=text)
                text = ''
            except Exception as err:
                au.video_url_file_id = ''
                au.save()
        if au.image or au.image_url_file_id:
            save_file_id = True

            if au.image_url_file_id:
                try:
                    await message.answer_photo(photo=au.image_url_file_id, caption=text)
                    text = ''
                    save_file_id = False
                except Exception as err:
                    au.image_url_file_id = ''
                    au.save()

            if save_file_id and au.image.name:
                try:
                    result_photo = await message.answer_photo(photo=await get_photo(au.image.name), caption=text)
                    text = ''
                    au.image_url_file_id = result_photo.photo[-1].file_id
                    au.save()
                except Exception as err:
                    au.image.name = ''
                    au.save()
        if text:
            await message.answer(text=text)
        if au.latitude and au.longitude:
            try:
                await message.answer_location(latitude=float(au.latitude), longitude=float(au.longitude))
            except Exception as err:
                au.latitude = ''
                au.longitude = ''
                au.save()


async def get_photo(url: str = None, file_id: str = None, path: str = None):
    if file_id:
        return file_id
    else:
        if not path:
            path = config.PATH_MEDIA
        return types.InputFile(path_or_bytesio=path + url)


async def welcome_message(message: types.Message, language: str):
    welcome_messages = await db_commands.get_welcomes_messages(language=language)
    for wm in welcome_messages:
        text = ''
        if wm.title:
            text += md.hbold(wm.title)
        if wm.description:
            text += '\n' + wm.description
        text = await replace_text_special(message=message, text=text)
        if wm.video_url_file_id:
            try:
                video_file_id = wm.video_url_file_id
                await message.answer_video(video=video_file_id, caption=text)
                text = ''
            except Exception as err:
                wm.video_url_file_id = ''
                wm.save()
        if wm.image or wm.image_url_file_id:
            save_file_id = True

            if wm.image_url_file_id:
                try:
                    await message.answer_photo(photo=wm.image_url_file_id, caption=text)
                    text = ''
                    save_file_id = False
                except Exception as err:
                    wm.image_url_file_id = ''
                    wm.save()

            if save_file_id and wm.image.name:
                try:
                    result_photo = await message.answer_photo(photo=await get_photo(wm.image.name), caption=text)
                    text = ''
                    wm.image_url_file_id = result_photo.photo[-1].file_id
                    wm.save()
                except Exception as err:
                    wm.image.name = ''
                    wm.save()
        if text:
            await message.answer(text=text)
        if wm.latitude and wm.longitude:
            try:
                await message.answer_location(latitude=float(wm.latitude), longitude=float(wm.longitude))
            except Exception as err:
                wm.latitude = ''
                wm.longitude = ''
                wm.save()


async def replace_text_special(text: str, message: types.Message):
    text = text.replace('{user}', message.from_user.full_name)
    return text


async def after_item_to_basket(language: str, call: types.CallbackQuery, state: FSMContext, item_name: str, user: User):
    message = call.message
    await message.delete()

    edit_text = f"<b>{item_name}</b>\n\n{values_title.ENTER_QUANTITY[language]}"
    # if message.text:
    #     await message.edit_text(reply_markup=None, text=edit_text)
    # elif message.photo or message.video:
    #     await message.edit_caption(reply_markup=None, caption=edit_text)
    message_caption = await message.answer(text=edit_text, reply_markup=None)
    markup = kb.enter_quantity(language=language)
    message_edit = await message.answer("#️⃣", reply_markup=markup)
    await state.update_data(message_edit=message_edit, message_caption=message_caption)
    await StateDefault.buy_item.set()


async def after_enter_quantity(user: models.User, quantity: int, item_id_adding: str, language: str,
                               message: types.Message, message_caption: types.Message,
                               message_edit: types.Message):
    attr = others.get_attribute_by_language(language=language, name_object='name')

    item = await db_commands.get_item_filter(first=True, id=item_id_adding)
    if item is not None:
        quantity_after = 0
        item_name = getattr(item, attr)
        additional_order = await state_data.get_structure_from_more_order()
        access_more_order = additional_order['access']
        if access_more_order:
            over_the_limit_add_order = await botControlOrder.over_the_limit_add_order(user=user, item=item,
                                                                                      quantity=quantity)
            if over_the_limit_add_order:
                await botControlOrder.prohibited_add_order(user=user, message=message)
                return
            await state_data.update_item_to_more_order(item=item, quantity=quantity)
            quantity_after = quantity
        else:
            basket_items = await db_commands.get_or_create_basket(user=user, product=item)
            len_basket_items = len(basket_items)
            for basket_item in basket_items[:]:
                if basket_item == basket_items[len_basket_items - 1] and quantity != 0:
                    # basket_item.quantity = (basket_item.quantity if basket_item.quantity > 1 else 0) + quantity
                    basket_item.quantity = quantity
                    basket_item.save()

                    quantity_after = basket_item.quantity
                else:
                    basket_item.delete()  # clear duplicate products
                    quantity_after = 0
        successfully_text = values_title.get_recorded_to_basket(item_name=item_name, quantity=quantity_after,
                                                                price=item.price, language=language,
                                                                remainder=item.reminder,
                                                                access_more_order=access_more_order)

        markup = await inlines.add_to_basket(language=language, item_id=item.id, quantity=quantity_after)
        if message_caption.text:
            await message_caption.edit_text(reply_markup=markup, text=successfully_text)
        elif message_caption.caption:
            await message_caption.edit_caption(reply_markup=markup, caption=successfully_text)
        await show_orders(message=message, user=user)
        # await after_product(language=language, message=message, text=values_title.WANT_TO_ADD_BASKET[language])
    else:
        await message_caption.delete()
        await message.answer(values_title.NOT_FOUND_ITEM[language])
    if message_edit:
        await message_edit.delete()
    await message.delete()


async def reset_basket(language: str, user: User, message: types.Message):
    result = await db_commands.delete_all_basket_by_user(user=user)
    if result:
        await message.answer(values_title.CART_EMPTIED[language])
    else:
        await message.answer(values_title.NOT_FOUND_ITEM[language])
    await home_state(language=language, message=message)


async def review(language: str, message: types.Message):
    main_reviews = await menu_values.get_mainReviews(language=language, show_bot=True)
    if main_reviews:
        markup = await kb.reviews_keyboards(language=language, reviews=main_reviews)
        await message.answer(text=values_title.CHOOSE_CATEGORY_OF_REVIEW[language], reply_markup=markup)
        await StateDefault.review.set()
    else:
        await review_text(language=language, message=message)


async def review_text(language: str, message: types.Message):
    markup = await kb.back_keyboard(language=language)
    await message.answer(text=values_title.REVIEW_MESSAGE[language], reply_markup=markup)
    await StateDefault.review_text.set()


async def forward_to_channels(user: User, message: types.Message, text: str = '', send_client: bool = True):
    constant = Constants.default_constant()
    for channel in constant.channels_tg():
        try:
            if text:
                message_forwarded = await bot.send_message(chat_id=channel, text=text)
            else:
                message_forwarded = await message.forward(chat_id=channel)
            client = values_title.NOT_SET[user.language]
            if user.client:
                client = user.client.name
            if send_client:
                await message_forwarded.reply(text=f"Клиент: <b>{client}</b>")
        except:
            pass


async def edit_quantity_product(language: str, user: User, message: types.Message, state):
    basket_items = await db_commands.get_basket_item(user=user)
    await state.update_data(all_message_edit_quantity=[])
    if len(basket_items) == 0:
        await send_no_baskets(message=message, language=language)
        return
    all_message_edit_quantity = []
    attr = others.get_attribute_by_language(language=language)
    i = 0
    for basket_item in basket_items:
        item_name = getattr(basket_item.product, attr)
        text = others.get_basket_card(name_text=item_name, num=basket_item.product.price, language=language,
                                      remainder=basket_item.product.reminder, quantity=basket_item.quantity)
        # markup = await inlines.get_inline_kb_edit_item_basket(basket_id=basket_item.id,
        #                                                       quantity=basket_item.quantity,
        #                                                       language=language)
        markup = await inlines.add_to_basket(item_id=basket_item.product.id,
                                             quantity=basket_item.quantity, language=language)
        message_edit = await message.answer(text=text, reply_markup=markup)
        all_message_edit_quantity.append(message_edit)
        i = i + 1
        if i % 29 == 0:
            # await asyncio.sleep(1)
            pass
    await state.update_data(all_message_edit_quantity=all_message_edit_quantity)


async def send_item(item, user: User, message: types.Message, save_data: bool = True, photo_or_video: bool = True):
    language = user.language
    # ------------------------------- basket ----------------------
    structure_more_order = await state_data.get_structure_from_more_order()
    quantity = 0
    if structure_more_order['access']:  # more_order
        _structure_more_order = await state_data.get_item_from_more_order(item=item)
        quantity = _structure_more_order['quantity']
    else:
        basketItem = await db_commands.get_basket_item(first=True, user=user, product=item)
        if basketItem:
            quantity = basketItem.quantity

    markup = await inlines.add_to_basket(language=language, quantity=quantity, item_id=item.id)
    text = await values_title.get_text_for_item_2(quantity=quantity, item=item, language=language)
    message_caption = None
    if item.video_url_file_id and photo_or_video:
        try:
            video_file_id = item.video_url_file_id
            message_caption = await message.answer_video(video=video_file_id, caption=text, reply_markup=markup)
            text = ''
            markup = None
        except Exception as err:
            item.video_url_file_id = ''
            item.save()
    if (item.image or item.image_url_file_id) and photo_or_video:
        save_file_id = True

        if item.image_url_file_id:
            try:
                message_caption = await message.answer_photo(photo=item.image_url_file_id, caption=text,
                                                             reply_markup=markup)
                text = ''
                markup = None
                save_file_id = False
            except Exception as err:
                item.image_url_file_id = ''
                item.save()

        if save_file_id and item.image.name:
            try:
                result_photo = await message.answer_photo(photo=await get_photo(item.image.name), caption=text,
                                                          reply_markup=markup)
                text = ''
                markup = None
                item.image_url_file_id = result_photo.photo[-1].file_id
                item.save()

                message_caption = result_photo
            except Exception as err:
                item.image.name = ''
                item.save()
    if text:
        message_caption = await message.answer(text=text, reply_markup=markup)
    if message_caption and save_data:
        await dp.current_state().update_data(message_caption=message_caption, item_id_adding=item.id)


async def act_sverki(user: User, message: types.Message):
    language = user.language
    if user.client is None:
        await not_found_client(language=language, message=message)
        await home_state(language=language, message=message, reset=True)
        return
    markup = kb.period(language=language)
    await message.answer(values_title.def_determine_period(language=language), reply_markup=markup)
    await StateDefault.act_sverki.set()


async def not_found_client(language: str, message: types.Message):
    await message.answer(
        f'{values_title.NOT_FOUND_CLIENT[language]}\n\n{values_title.CONTACT_TO_REVIEW[language]}')


async def search(language: str, message: types.Message):
    markup = types.InlineKeyboardMarkup()
    markup.insert(types.InlineKeyboardButton(text=menu_values.SEARCH[language],
                                             switch_inline_query_current_chat=""))
    message = await message.answer(values_title.GO_SEARCH[language], reply_markup=markup)
    return message


async def reset_data_user():
    data = await dp.current_state().get_data()
    basket_message = data.get("basket_message")
    client_chosen = data.get("client_chosen")
    await dp.current_state().reset_data()

    await dp.current_state().update_data(basket_message=basket_message, client_chosen=client_chosen)


async def send_to_group(text: str, user: User = None):
    constant = Constants.default_constant()
    for channel in constant.channels_tg():
        try:
            await dp.bot.send_message(chat_id=channel, text=(text if text == text[:200] else f"{text[:200]}.."))
        except:
            pass


def send_to_group_sync(text: str, user: User = None):
    constant = Constants.default_constant()
    for channel in constant.channels_tg():
        try:
            bot_tel.send_message(chat_id=channel, text=(text if text == text[:300] else f"{text[:300]}.."))
        except:
            pass


async def send_file_items(message: types.Message, language: str):
    # try:
    await message.answer(values_title.LOADING[language])
    items = await db_commands.get_item_filter(show_bot=True, running=True)
    if len(items) == 0 or items is None:
        await message.answer(values_title.NOT_FOUND_ITEM[language])
        return
    import openpyxl
    import os
    from openpyxl import styles as st
    import tempfile

    current_date = dt.datetime.now()
    # output_file = 'Дефицитный лист.xlsx'
    tmp_xlsx = tempfile.NamedTemporaryFile(suffix='.xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active

    # ----------------------------------------- STYLES -----------------------------------
    name_font = 'Arial'
    side_thin = st.Side(style='thin', color='000000')
    font_10_bold = st.Font(name=name_font, sz=10, b=True)
    font_12_bold = st.Font(name=name_font, sz=12, b=True)
    font_11_bold = st.Font(name=name_font, sz=11, b=True, color='FFFFFF')
    align_center = st.Alignment(horizontal='center',
                                vertical='center',
                                wrapText=True)
    left_center = st.Alignment(horizontal='left', vertical='center', wrapText=True)
    border_all_thin = st.Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin)

    # -----------------------------------------------------------------------------------

    # -------------------------------- RESIZE COLUMNS AND ROW----------------------------
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    # ws.column_dimensions['C'].width = 10
    # ws.column_dimensions['D'].width = 15
    # ws.column_dimensions['E'].width = 10
    # ----------------------------------------------------------------------------
    columns = [{'B': 'name'}]
    # ----------------------------- playing with cells -------------------------
    ws.row_dimensions[1].height = 30.0
    ws.row_dimensions[2].height = 9.0
    ws.row_dimensions[3].height = 25.5

    ws[
        'A1'] = f"Дефицитный лист от {current_date.strftime('%d.%m.%Y')} на время {int(current_date.time().hour)}:{current_date.strftime('%M:%S')}"
    ws['A1'].font = font_12_bold
    ws['A1'].alignment = align_center
    ws.merge_cells('A1:B1')

    num = 'A3'
    ws[num] = '№'
    ws[num].font = font_11_bold
    ws[num].alignment = align_center

    num = 'B3'
    ws[num] = 'Товар'
    ws[num].font = font_11_bold
    ws[num].alignment = align_center

    # num = 'C3'
    # ws[num] = 'Остаток'
    # ws[num].font = font_11_bold
    # ws[num].alignment = align_center
    #
    # num = 'D3'
    # ws[num] = 'Дефицитный'
    # ws[num].font = font_11_bold
    # ws[num].alignment = align_center

    # num = 'E3'
    # ws[num] = 'Нагрузка'
    # ws[num].font = font_11_bold
    # ws[num].alignment = align_center

    num = 1
    for item in items:
        cell_name = f'A{num + 3}'
        ws[cell_name] = num
        ws[cell_name].alignment = left_center
        ws[cell_name].font = font_10_bold

        for data in columns:
            key_data, value_data = tuple(data.items())[0]
            cell_name = f'{key_data}{num + 3}'
            try:
                value = getattr(item, value_data)
                if isinstance(value, bool):
                    value = '+' if value else '-'
                ws[cell_name] = value
            except Exception as err:
                logging.info(err)
            ws[cell_name].alignment = left_center if key_data == 'B' else align_center
            ws[cell_name].font = font_10_bold

        num = num + 1

    for row in ws['A3:B3']:
        for cell in row:
            cell.border = border_all_thin
            cell.fill = st.PatternFill("solid", fgColor="4A82BD")

    cell_range = f"A4:B{ws.max_row}"
    i = 0
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border_all_thin
            if i % 2 == 0:
                cell.fill = st.PatternFill("solid", fgColor="82D3FF")
        i = i + 1

    wb.save(tmp_xlsx.name)
    if config.CONVERT_FILE_TO_PNG:
        import jpype
        import asposecells
        jpype.startJVM()
        from asposecells.api import Workbook, SheetRender, ImageOrPrintOptions, SaveFormat

        # load the Excel workbook
        workbook = Workbook(tmp_xlsx.name)

        # create image options
        imgOptions = ImageOrPrintOptions()
        imgOptions.setHorizontalResolution(200)
        imgOptions.setVerticalResolution(200)
        # imgOptions.setSaveFormat(SaveFormat.JPG)

        # load the worksheet to be rendered
        sheet = workbook.getWorksheets().get(0)

        # create sheet render object
        sr = SheetRender(sheet, imgOptions)

        # convert sheet to PNG image
        for j in range(0, sr.getPageCount()):
            fp = tempfile.NamedTemporaryFile(suffix='.png')
            sr.toImage(j, fp.name)
            await message.answer_document(document=types.InputFile(fp.name))
            fp.close()

        jpype.shutdownJVM()
    else:
        await message.answer_document(document=types.InputFile(tmp_xlsx.name))

    tmp_xlsx.close()


# except Exception as err:
#     logging.info(err)


async def not_allowed_to_act_sverki(user: User, message: types.Message):
    await message.answer(values_title.NOT_ACCESS_TO_ACT_SVERKI[user.language])
    await home_state(language=user.language, message=message)
